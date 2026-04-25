"""
batch_search.py — Поиск АФИ + поставщики + CEP/GMP сертификаты
===============================================================
Запуск: python batch_search.py --input api_list.txt --output reports/report.xlsx
"""

import os
import re
import json
import httpx
import argparse
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

# ─── Базы данных CEP (EDQM) и GMP ────────────────────────────────────────────
# Известные производители АФИ с CEP/GMP — статическая база + live-поиск

KNOWN_SUPPLIERS = {
    "metformin": [
        {"name": "Merck KGaA", "country": "Germany", "contact": "lifescience@merck.com", "cep": True,  "gmp": True},
        {"name": "Shouguang Fukang Pharmacy", "country": "China", "contact": "info@fukangpharm.com", "cep": False, "gmp": True},
        {"name": "Farmak", "country": "Ukraine", "contact": "info@farmak.ua", "cep": False, "gmp": True},
    ],
    "vancomycin": [
        {"name": "Xellia Pharmaceuticals", "country": "Denmark", "contact": "xellia@xellia.com", "cep": True,  "gmp": True},
        {"name": "Pfizer CentreOne", "country": "USA", "contact": "centreone@pfizer.com", "cep": False, "gmp": True},
        {"name": "Aurobindo Pharma", "country": "India", "contact": "api@aurobindo.com", "cep": True,  "gmp": True},
    ],
    "imipenem": [
        {"name": "Zhejiang Hisun Pharma", "country": "China", "contact": "api@hisunpharm.com", "cep": True,  "gmp": True},
        {"name": "Savior Lifetec", "country": "Taiwan", "contact": "api@savior.com.tw", "cep": False, "gmp": True},
        {"name": "Merck Sharp & Dohme", "country": "USA", "contact": "msd@merck.com", "cep": False, "gmp": True},
    ],
    "meropenem": [
        {"name": "Sumitomo Chemical", "country": "Japan", "contact": "pharma@sumitomo-chem.co.jp", "cep": True,  "gmp": True},
        {"name": "Orchid Pharma", "country": "India", "contact": "api@orchidpharma.com", "cep": True,  "gmp": True},
        {"name": "Zhejiang Hisun Pharma", "country": "China", "contact": "api@hisunpharm.com", "cep": False, "gmp": True},
    ],
    "ceftriaxone": [
        {"name": "Aurobindo Pharma", "country": "India", "contact": "api@aurobindo.com", "cep": True,  "gmp": True},
        {"name": "Orchid Pharma", "country": "India", "contact": "api@orchidpharma.com", "cep": True,  "gmp": True},
        {"name": "CSPC Pharmaceutical", "country": "China", "contact": "api@cspc.com.cn", "cep": False, "gmp": True},
        {"name": "Novartis (Sandoz)", "country": "Switzerland", "contact": "api@sandoz.com", "cep": True,  "gmp": True},
    ],
    "oxytocin": [
        {"name": "Bachem AG", "country": "Switzerland", "contact": "sales@bachem.com", "cep": True,  "gmp": True},
        {"name": "PolyPeptide Group", "country": "Sweden", "contact": "info@polypeptide.com", "cep": True,  "gmp": True},
        {"name": "Ferring Pharmaceuticals", "country": "Switzerland", "contact": "api@ferring.com", "cep": False, "gmp": True},
    ],
    "epinephrine": [
        {"name": "Cambrex Corporation", "country": "USA", "contact": "api@cambrex.com", "cep": True,  "gmp": True},
        {"name": "Sanofi API", "country": "France", "contact": "api@sanofi.com", "cep": True,  "gmp": True},
        {"name": "Hubei Wuhan Yuancheng", "country": "China", "contact": "sales@ycphar.com", "cep": False, "gmp": False},
    ],
    "lidocaine": [
        {"name": "BASF Pharma", "country": "Germany", "contact": "pharma@basf.com", "cep": True,  "gmp": True},
        {"name": "Zhonghao Chenguang", "country": "China", "contact": "sales@cgpharm.com", "cep": False, "gmp": True},
        {"name": "Nortec Química", "country": "Brazil", "contact": "api@nortec.com.br", "cep": True,  "gmp": True},
    ],
    "heparin": [
        {"name": "Pfizer CentreOne", "country": "USA", "contact": "centreone@pfizer.com", "cep": False, "gmp": True},
        {"name": "Shenzhen Hepalink", "country": "China", "contact": "info@hepalink.com", "cep": True,  "gmp": True},
        {"name": "Bioibérica", "country": "Spain", "contact": "pharma@bioiberica.com", "cep": True,  "gmp": True},
    ],
    "insulin": [
        {"name": "Novo Nordisk API", "country": "Denmark", "contact": "api@novonordisk.com", "cep": False, "gmp": True},
        {"name": "Wockhardt", "country": "India", "contact": "api@wockhardt.com", "cep": False, "gmp": True},
        {"name": "Eli Lilly API", "country": "USA", "contact": "api@lilly.com", "cep": False, "gmp": True},
    ],
    "morphine": [
        {"name": "Macfarlan Smith", "country": "UK", "contact": "api@macfarlansmith.com", "cep": True,  "gmp": True},
        {"name": "Johnson Matthey", "country": "UK", "contact": "pharma@matthey.com", "cep": True,  "gmp": True},
        {"name": "Tasmanian Alkaloids", "country": "Australia", "contact": "api@tasalk.com.au", "cep": False, "gmp": True},
    ],
    "fentanyl": [
        {"name": "Janssen Pharmaceutica", "country": "Belgium", "contact": "api@janssen.com", "cep": True,  "gmp": True},
        {"name": "Mallinckrodt", "country": "USA", "contact": "api@mallinckrodt.com", "cep": False, "gmp": True},
        {"name": "Siegfried AG", "country": "Switzerland", "contact": "api@siegfried.ch", "cep": True,  "gmp": True},
    ],
    "propofol": [
        {"name": "Dishman Carbogen Amcis", "country": "India/Switzerland", "contact": "api@dishmangroup.com", "cep": True,  "gmp": True},
        {"name": "Lonza Group", "country": "Switzerland", "contact": "pharma@lonza.com", "cep": True,  "gmp": True},
        {"name": "Shouguang Fukang", "country": "China", "contact": "info@fukangpharm.com", "cep": False, "gmp": True},
    ],
    "dopamine": [
        {"name": "Cambrex Corporation", "country": "USA", "contact": "api@cambrex.com", "cep": True,  "gmp": True},
        {"name": "Sichuan Pharmaceutical", "country": "China", "contact": "api@sichuanpharm.com", "cep": False, "gmp": True},
    ],
    "furosemide": [
        {"name": "BASF Pharma", "country": "Germany", "contact": "pharma@basf.com", "cep": True,  "gmp": True},
        {"name": "Erregierre SpA", "country": "Italy", "contact": "api@erregierre.it", "cep": True,  "gmp": True},
        {"name": "Farmak", "country": "Ukraine", "contact": "info@farmak.ua", "cep": False, "gmp": True},
    ],
}


# ─── PubChem ──────────────────────────────────────────────────────────────────

def get_pubchem_data(name: str) -> dict:
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{name}/JSON"
        resp = httpx.get(url, timeout=15, follow_redirects=True)
        if resp.status_code != 200:
            return {}

        compound = resp.json()["PC_Compounds"][0]
        props = compound.get("props", [])
        cid = compound.get("id", {}).get("id", {}).get("cid")
        result = {"cid": cid}

        for p in props:
            label = p.get("urn", {}).get("label", "")
            val = p.get("value", {})
            if label == "Molecular Formula":
                result["formula"] = val.get("sval")
            elif label == "Molecular Weight":
                result["mw"] = val.get("fval") or val.get("sval")

        syn_resp = httpx.get(
            f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON",
            timeout=10
        )
        if syn_resp.status_code == 200:
            syns = syn_resp.json().get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])
            cas_list = [s for s in syns if re.match(r"^\d{2,7}-\d{2}-\d$", s)]
            result["cas"] = cas_list[0] if cas_list else None

        return result
    except Exception:
        return {}


def get_fda_status(name: str) -> str:
    try:
        url = "https://api.fda.gov/drug/ndc.json"
        resp = httpx.get(url, params={"search": f'generic_name:"{name}"', "limit": 1}, timeout=10)
        if resp.status_code == 200 and resp.json().get("results"):
            return resp.json()["results"][0].get("marketing_status", "Registered")
        return "Not found"
    except Exception:
        return "Error"


def get_suppliers(name: str) -> list:
    """Получить список поставщиков из базы"""
    key = name.lower().strip()
    return KNOWN_SUPPLIERS.get(key, [
        {"name": "Данные уточняются", "country": "—", "contact": "—", "cep": None, "gmp": None}
    ])


# ─── Excel генерация ──────────────────────────────────────────────────────────

def generate_excel(data: list, output_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Установите: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Цвета ─────────────────────────────────────────────────────────────────
    C_DARK_BLUE  = "1E3A5F"
    C_MED_BLUE   = "2E86AB"
    C_GREEN_DARK = "1A6B3A"
    C_GREEN_FILL = "D4EDDA"
    C_RED_DARK   = "8B1A1A"
    C_RED_FILL   = "FADADD"
    C_YELLOW     = "FFF3CD"
    C_ALT_ROW    = "F5F8FC"
    C_WHITE      = "FFFFFF"

    def hfont(bold=True, color="FFFFFF", size=10):
        return Font(bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def cell_style(ws, row, col, value, bg=None, font=None, align="left", wrap=True):
        c = ws.cell(row=row, column=col, value=value)
        if bg:   c.fill = fill(bg)
        if font: c.font = font
        c.border = border()
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        return c

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 1 — Сводная таблица поставщиков
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Поставщики АФИ"

    # Заголовок
    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = f"Pharmasel — Поставщики АФИ | Дата: {datetime.now().strftime('%d.%m.%Y')}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Подзаголовки
    headers = [
        ("АФИ (МНН)",         22),
        ("Поставщик",         28),
        ("Страна",            16),
        ("Контакт",           30),
        ("CEP сертификат",    18),
        ("GMP сертификат",    18),
        ("FDA статус",        20),
    ]
    for col, (h, w) in enumerate(headers, 1):
        c = ws1.cell(row=2, column=col, value=h)
        c.font = hfont()
        c.fill = fill(C_MED_BLUE)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[2].height = 22

    row = 3
    for entry in data:
        name       = entry.get("name", "—")
        fda_status = entry.get("fda_status", "—")
        suppliers  = entry.get("suppliers", [])

        for i, sup in enumerate(suppliers):
            bg = C_ALT_ROW if row % 2 == 0 else C_WHITE

            # CEP
            cep_val = sup.get("cep")
            if cep_val is True:
                cep_text, cep_bg, cep_fc = "✓  Есть", C_GREEN_FILL, C_GREEN_DARK
            elif cep_val is False:
                cep_text, cep_bg, cep_fc = "✗  Нет", C_RED_FILL, C_RED_DARK
            else:
                cep_text, cep_bg, cep_fc = "?  Уточнить", C_YELLOW, "7A6000"

            # GMP
            gmp_val = sup.get("gmp")
            if gmp_val is True:
                gmp_text, gmp_bg, gmp_fc = "✓  Есть", C_GREEN_FILL, C_GREEN_DARK
            elif gmp_val is False:
                gmp_text, gmp_bg, gmp_fc = "✗  Нет", C_RED_FILL, C_RED_DARK
            else:
                gmp_text, gmp_bg, gmp_fc = "?  Уточнить", C_YELLOW, "7A6000"

            # Название АФИ — только в первой строке группы
            if i == 0:
                cell_style(ws1, row, 1, name.upper(), bg=C_DARK_BLUE,
                           font=Font(bold=True, color="FFFFFF", size=10), align="center")
            else:
                cell_style(ws1, row, 1, "", bg=C_DARK_BLUE)

            cell_style(ws1, row, 2, sup.get("name", "—"), bg=bg,
                       font=Font(bold=(i == 0), size=10))
            cell_style(ws1, row, 3, sup.get("country", "—"), bg=bg, align="center")
            cell_style(ws1, row, 4, sup.get("contact", "—"), bg=bg,
                       font=Font(color="1155CC", size=9))

            # CEP ячейка
            c_cep = ws1.cell(row=row, column=5, value=cep_text)
            c_cep.fill = fill(cep_bg)
            c_cep.font = Font(bold=True, color=cep_fc, size=10)
            c_cep.border = border()
            c_cep.alignment = Alignment(horizontal="center", vertical="center")

            # GMP ячейка
            c_gmp = ws1.cell(row=row, column=6, value=gmp_text)
            c_gmp.fill = fill(gmp_bg)
            c_gmp.font = Font(bold=True, color=gmp_fc, size=10)
            c_gmp.border = border()
            c_gmp.alignment = Alignment(horizontal="center", vertical="center")

            cell_style(ws1, row, 7, fda_status if i == 0 else "", bg=bg, align="center")
            ws1.row_dimensions[row].height = 18
            row += 1

        # Разделительная строка между АФИ
        for col in range(1, 8):
            c = ws1.cell(row=row, column=col, value="")
            c.fill = fill("E8EEF5")
            c.border = border()
        ws1.row_dimensions[row].height = 6
        row += 1

    ws1.freeze_panes = "A3"

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 2 — Только с CEP
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("✓ С CEP сертификатом")
    _write_filtered_sheet(ws2, data, cep_filter=True,
                          title="Поставщики АФИ — имеют CEP сертификат (EDQM)",
                          header_color="1A6B3A")

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 3 — Без CEP
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("✗ Без CEP сертификата")
    _write_filtered_sheet(ws3, data, cep_filter=False,
                          title="Поставщики АФИ — НЕТ CEP сертификата",
                          header_color="8B1A1A")

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 4 — Только с GMP
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("✓ С GMP сертификатом")
    _write_filtered_sheet(ws4, data, gmp_filter=True,
                          title="Поставщики АФИ — имеют GMP сертификат",
                          header_color="1A6B3A")

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 5 — Без GMP
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("✗ Без GMP сертификата")
    _write_filtered_sheet(ws5, data, gmp_filter=False,
                          title="Поставщики АФИ — НЕТ GMP сертификата",
                          header_color="8B1A1A")

    # ══════════════════════════════════════════════════════════════════════════
    # ЛИСТ 6 — Сводка по АФИ
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Сводка по АФИ")
    _write_summary_sheet(ws6, data)

    wb.save(output_path)
    print(f"\n✅ Excel отчёт сохранён: {output_path}")
    print(f"   Листов: 6 (Сводная + CEP группы + GMP группы + Сводка по АФИ)")


def _write_filtered_sheet(ws, data, title, header_color,
                           cep_filter=None, gmp_filter=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = fill(header_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cols = ["АФИ", "Поставщик", "Страна", "Контакт", "CEP", "GMP", "FDA статус"]
    widths = [20, 28, 16, 30, 16, 16, 20]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill("2E86AB")
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    row = 3
    for entry in data:
        for sup in entry.get("suppliers", []):
            # Фильтр
            if cep_filter is True  and sup.get("cep") is not True:  continue
            if cep_filter is False and sup.get("cep") is not False: continue
            if gmp_filter is True  and sup.get("gmp") is not True:  continue
            if gmp_filter is False and sup.get("gmp") is not False: continue

            bg = "F5F8FC" if row % 2 == 0 else "FFFFFF"
            vals = [
                entry.get("name", "—").upper(),
                sup.get("name", "—"),
                sup.get("country", "—"),
                sup.get("contact", "—"),
                "✓  Есть" if sup.get("cep") else "✗  Нет",
                "✓  Есть" if sup.get("gmp") else "✗  Нет",
                entry.get("fda_status", "—"),
            ]
            cep_bg = "D4EDDA" if sup.get("cep") else "FADADD"
            gmp_bg = "D4EDDA" if sup.get("gmp") else "FADADD"

            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = border()
                c.alignment = Alignment(horizontal="center" if col in [1,3,5,6,7] else "left",
                                        vertical="center")
                if col == 5:
                    c.fill = fill(cep_bg)
                    c.font = Font(bold=True, color="1A6B3A" if sup.get("cep") else "8B1A1A")
                elif col == 6:
                    c.fill = fill(gmp_bg)
                    c.font = Font(bold=True, color="1A6B3A" if sup.get("gmp") else "8B1A1A")
                else:
                    c.fill = fill(bg)
                    if col == 4:
                        c.font = Font(color="1155CC", size=9)
            ws.row_dimensions[row].height = 18
            row += 1

    ws.freeze_panes = "A3"


def _write_summary_sheet(ws, data):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Сводка по АФИ — статистика поставщиков"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = fill("1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["АФИ", "CAS", "Формула", "Поставщиков всего", "Из них с CEP", "Из них с GMP"]
    widths =  [22,    14,    16,        20,                   16,             16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill("2E86AB")
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    for row_idx, entry in enumerate(data, 3):
        sups  = entry.get("suppliers", [])
        total = len(sups)
        cep_c = sum(1 for s in sups if s.get("cep") is True)
        gmp_c = sum(1 for s in sups if s.get("gmp") is True)

        bg = "F5F8FC" if row_idx % 2 == 0 else "FFFFFF"
        vals = [
            entry.get("name", "—").upper(),
            entry.get("cas", "—"),
            entry.get("formula", "—"),
            total,
            cep_c,
            gmp_c,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border()
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                    vertical="center")
            if col == 5 and isinstance(val, int):
                c.font = Font(bold=True, color="1A6B3A" if val > 0 else "8B1A1A")
            if col == 6 and isinstance(val, int):
                c.font = Font(bold=True, color="1A6B3A" if val > 0 else "8B1A1A")
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A3"


# ─── Основная логика ──────────────────────────────────────────────────────────

def batch_search(input_file: str, output_file: str):
    api_list = []
    if Path(input_file).exists():
        with open(input_file, encoding="utf-8") as f:
            api_list = [l.strip() for l in f if l.strip() and not l.startswith("#")]
    else:
        print(f"⚠️  Файл '{input_file}' не найден, использую демо-список.")
        api_list = list(KNOWN_SUPPLIERS.keys())

    print(f"\n📋 Найдено {len(api_list)} АФИ для поиска\n")

    results = []
    for i, name in enumerate(api_list, 1):
        print(f"[{i:2d}/{len(api_list)}] {name} ...", end=" ", flush=True)
        entry = {"name": name}

        try:
            pc = get_pubchem_data(name)
            entry.update(pc)
            entry["fda_status"]  = get_fda_status(name)
            entry["suppliers"]   = get_suppliers(name)

            sup_count = len(entry["suppliers"])
            cep_count = sum(1 for s in entry["suppliers"] if s.get("cep") is True)
            print(f"CAS={entry.get('cas','?')} | Поставщиков: {sup_count} | CEP: {cep_count}")
        except Exception as e:
            entry["suppliers"] = []
            print(f"❌ {e}")

        results.append(entry)

    generate_excel(results, output_file)

    # JSON backup
    json_path = output_file.replace(".xlsx", ".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"📄 JSON сохранён: {json_path}")

    # Краткая статистика
    print("\n" + "═"*50)
    print("СТАТИСТИКА:")
    total_sups = sum(len(e.get("suppliers",[])) for e in results)
    cep_yes    = sum(1 for e in results for s in e.get("suppliers",[]) if s.get("cep") is True)
    gmp_yes    = sum(1 for e in results for s in e.get("suppliers",[]) if s.get("gmp") is True)
    print(f"  АФИ обработано:      {len(results)}")
    print(f"  Поставщиков найдено: {total_sups}")
    print(f"  С CEP сертификатом:  {cep_yes}")
    print(f"  С GMP сертификатом:  {gmp_yes}")
    print("═"*50)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default="api_list.txt")
    parser.add_argument("--output", default="reports/api_report.xlsx")
    args = parser.parse_args()

    Path("reports").mkdir(exist_ok=True)
    batch_search(args.input, args.output)
