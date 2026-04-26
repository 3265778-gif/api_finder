"""
api_finder.py — Поиск поставщиков АФИ / сырья для БАДов
========================================================
Два режима:
  1. Фарма (АФИ) — проверка CEP + GMP
  2. БАД (сырьё)  — проверка ISO 22000 + FSSC 22000

Запуск: python api_finder.py
"""

import os
import re
import json
import httpx
import time
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)


def parse_input(text):
    results = []
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    for line in lines:
        line = line.lstrip("*-. ")
        cas_match = re.search(r"(\d{2,7}-\d{2}-\d)", line)
        cas = cas_match.group(1) if cas_match else None
        name = re.sub(r"\(?\s*CAS\s*:?\s*\d{2,7}-\d{2}-\d\s*\)?", "", line, flags=re.IGNORECASE)
        name = re.sub(r",?\s*\d{2,7}-\d{2}-\d", "", name).strip(" ,()-")
        if cas or name:
            results.append({"name": name or "Unknown", "cas": cas})
    return results


def pubchem_by_cas(cas):
    try:
        resp = httpx.get(f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{cas}/JSON",
                         timeout=15, follow_redirects=True)
        if resp.status_code != 200: return {}
        compound = resp.json()["PC_Compounds"][0]
        props = compound.get("props", [])
        cid = compound.get("id", {}).get("id", {}).get("cid")
        result = {"cid": cid}
        for p in props:
            label = p.get("urn", {}).get("label", "")
            val = p.get("value", {})
            if label == "Molecular Formula": result["formula"] = val.get("sval")
            elif label == "Molecular Weight": result["mw"] = val.get("fval") or val.get("sval")
            elif label == "IUPAC Name" and p.get("urn", {}).get("name") == "Preferred":
                result["iupac"] = val.get("sval")
        return result
    except: return {}


def pubchem_by_name(name):
    try:
        resp = httpx.get(f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{name}/JSON",
                         timeout=15, follow_redirects=True)
        if resp.status_code != 200: return {}
        compound = resp.json()["PC_Compounds"][0]
        props = compound.get("props", [])
        cid = compound.get("id", {}).get("id", {}).get("cid")
        result = {"cid": cid}
        for p in props:
            label = p.get("urn", {}).get("label", "")
            val = p.get("value", {})
            if label == "Molecular Formula": result["formula"] = val.get("sval")
            elif label == "Molecular Weight": result["mw"] = val.get("fval") or val.get("sval")
        syn_resp = httpx.get(
            f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON", timeout=10)
        if syn_resp.status_code == 200:
            syns = syn_resp.json().get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])
            cas_list = [s for s in syns if re.match(r"^\d{2,7}-\d{2}-\d$", s)]
            result["cas"] = cas_list[0] if cas_list else None
        return result
    except: return {}


def _normalize_cert(value):
    """Нормализует значение сертификата в bool или None.
    Claude API иногда возвращает 'true'/'false' как строку вместо boolean.
    Эта функция гарантирует что значение всегда True, False или None."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        if value.lower() in ("true", "yes", "1"):
            return True
        if value.lower() in ("false", "no", "0"):
            return False
    if isinstance(value, (int, float)):
        return bool(value)
    return None


def ai_search_suppliers(name, cas, formula, category):
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("    !! ANTHROPIC_API_KEY not set")
        return [{"name": "Supplier unknown", "country": "-", "contact": "-",
                 "cep": None, "gmp": None, "iso22000": None, "fssc22000": None}]

    if category == "bad":
        cert_part = """For EACH supplier provide:
1. Full company name
2. Country
3. Contact email or website
4. Does the company have ISO 22000 certificate (food safety) - true/false/null
5. Does the company have GMP certificate (optional for supplements) - true/false/null
6. Does the company have FSSC 22000 certificate - true/false/null

IMPORTANT: Use JSON boolean values true/false (NOT strings "true"/"false").

Reply with ONLY a valid JSON array:
[
  {"name": "Company", "country": "Country", "contact": "email@example.com", "iso22000": true, "gmp": true, "fssc22000": true}
]"""
        context = "dietary supplement ingredient / raw material"
    else:
        cert_part = """For EACH supplier provide:
1. Full company name
2. Country
3. Contact email or website
4. Does the company have CEP certificate (Certificate of Suitability EDQM) - true/false/null
5. Does the company have GMP certificate - true/false/null

IMPORTANT: Use JSON boolean values true/false (NOT strings "true"/"false").

Reply with ONLY a valid JSON array:
[
  {"name": "Company", "country": "Country", "contact": "email@example.com", "cep": true, "gmp": true}
]"""
        context = "active pharmaceutical ingredient (API)"

    prompt = f"""Find 3-5 real manufacturers/suppliers of {context}:
Name: {name}
CAS number: {cas}
Formula: {formula}

{cert_part}

If you are not sure about a certificate - use null.
Return ONLY valid JSON, nothing else. No markdown, no explanation."""

    try:
        resp = httpx.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 1500,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        if resp.status_code != 200:
            print(f"    !! Claude API error {resp.status_code}")
            return [{"name": "API error", "country": "-", "contact": "-",
                     "cep": None, "gmp": None, "iso22000": None, "fssc22000": None}]

        data = resp.json()
        text = ""
        for block in data.get("content", []):
            if block.get("type") == "text":
                text += block.get("text", "")
        text = text.strip()
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        suppliers = json.loads(text)
        if isinstance(suppliers, list):
            # Нормализуем все значения сертификатов
            for sup in suppliers:
                for key in ["cep", "gmp", "iso22000", "fssc22000"]:
                    if key in sup:
                        sup[key] = _normalize_cert(sup[key])
            return suppliers
    except json.JSONDecodeError:
        print("    !! JSON parse error")
    except Exception as e:
        print(f"    !! Error: {e}")

    return [{"name": "Supplier unknown", "country": "-", "contact": "-",
             "cep": None, "gmp": None, "iso22000": None, "fssc22000": None}]


def generate_filename(substances, category):
    names = [s["name"] for s in substances if s.get("name")]
    ts = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    prefix = "BAD" if category == "bad" else "API"
    def clean(n):
        n = re.sub(r"[^\w\s-]", "", n).strip()
        n = re.sub(r"\s+", "-", n)
        return n[:25]
    if len(names) == 0:   return f"{prefix}_search_{ts}.xlsx"
    elif len(names) == 1: return f"{prefix}_{clean(names[0])}_{ts}.xlsx"
    elif len(names) == 2: return f"{prefix}_{clean(names[0])}_{clean(names[1])}_{ts}.xlsx"
    else:
        return f"{prefix}_{clean(names[0])}_{clean(names[1])}_+{len(names)-2}_{ts}.xlsx"


def generate_excel(data, output_path, category):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("!! pip install openpyxl"); return

    wb = openpyxl.Workbook()
    C_DARK="1E3A5F"; C_BLUE="2E86AB"; C_GREEN_D="1A6B3A"; C_GREEN_F="D4EDDA"
    C_RED_D="8B1A1A"; C_RED_F="FADADD"; C_YELLOW="FFF3CD"; C_ALT="F5F8FC"; C_WHITE="FFFFFF"
    C_BAD_DARK="2C5F2D"; C_BAD_HEAD="3E8E41"

    def fl(c): return PatternFill("solid", fgColor=c)
    def brd():
        s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)
    def al(h="left"): return Alignment(horizontal=h,vertical="center",wrap_text=True)
    def cert_style(val):
        if val is True:   return "YES", C_GREEN_F, C_GREEN_D
        elif val is False: return "NO",  C_RED_F,   C_RED_D
        else:              return "?  Unknown", C_YELLOW, "7A6000"

    is_bad = category == "bad"
    ws = wb.active

    if is_bad:
        ws.title = "All suppliers (BAD)"
        title_text = f"Supplement ingredient suppliers | {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        headers = ["Ingredient","CAS","Supplier","Country","Contact","ISO 22000","GMP","FSSC 22000","Formula"]
        widths  = [24,14,28,14,30,16,16,16,18]
    else:
        ws.title = "All suppliers (Pharma)"
        title_text = f"API suppliers | {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        headers = ["API","CAS","Supplier","Country","Contact","CEP cert.","GMP cert.","Formula"]
        widths  = [24,14,28,14,30,16,16,18]

    nc = len(headers)
    hdr_color = C_BAD_DARK if is_bad else C_DARK
    sub_color = C_BAD_HEAD if is_bad else C_BLUE

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nc)
    c=ws.cell(row=1,column=1,value=title_text)
    c.font=Font(bold=True,size=13,color="FFFFFF"); c.fill=fl(hdr_color); c.alignment=al("center")
    ws.row_dimensions[1].height=30

    for i,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=2,column=i,value=h)
        c.font=Font(bold=True,color="FFFFFF",size=10); c.fill=fl(sub_color)
        c.border=brd(); c.alignment=al("center")
        ws.column_dimensions[get_column_letter(i)].width=w

    row=3
    for entry in data:
        sups=entry.get("suppliers",[])
        for j,sup in enumerate(sups):
            bg=C_ALT if row%2==0 else C_WHITE
            if is_bad:
                iso_t,iso_bg,iso_fc=cert_style(sup.get("iso22000"))
                gmp_t,gmp_bg,gmp_fc=cert_style(sup.get("gmp"))
                fssc_t,fssc_bg,fssc_fc=cert_style(sup.get("fssc22000"))
                vals=[entry["name"].title() if j==0 else "",
                      entry.get("cas","-") if j==0 else "",
                      sup.get("name","-"), sup.get("country","-"), sup.get("contact","-"),
                      iso_t, gmp_t, fssc_t,
                      entry.get("formula","-") if j==0 else ""]
                cert_cols={6:(iso_bg,iso_fc),7:(gmp_bg,gmp_fc),8:(fssc_bg,fssc_fc)}
            else:
                cep_t,cep_bg,cep_fc=cert_style(sup.get("cep"))
                gmp_t,gmp_bg,gmp_fc=cert_style(sup.get("gmp"))
                vals=[entry["name"].title() if j==0 else "",
                      entry.get("cas","-") if j==0 else "",
                      sup.get("name","-"), sup.get("country","-"), sup.get("contact","-"),
                      cep_t, gmp_t,
                      entry.get("formula","-") if j==0 else ""]
                cert_cols={6:(cep_bg,cep_fc),7:(gmp_bg,gmp_fc)}

            for col,val in enumerate(vals,1):
                c=ws.cell(row=row,column=col,value=val); c.border=brd()
                if col==1 and j==0:
                    c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=fl(hdr_color); c.alignment=al("center")
                elif col==1:
                    c.fill=fl(hdr_color)
                elif col==5:
                    c.font=Font(color="1155CC",size=9); c.fill=fl(bg); c.alignment=al()
                elif col in cert_cols:
                    cbg,cfc=cert_cols[col]; c.fill=fl(cbg); c.font=Font(bold=True,color=cfc); c.alignment=al("center")
                else:
                    c.fill=fl(bg); c.alignment=al("center") if col in [2,4,nc] else al()
            ws.row_dimensions[row].height=18; row+=1

        for col in range(1,nc+1):
            c=ws.cell(row=row,column=col); c.fill=fl("E5F0E5" if is_bad else "E8EEF5"); c.border=brd()
        ws.row_dimensions[row].height=5; row+=1

    ws.freeze_panes="A3"

    if is_bad:
        _cert_sheet(wb.create_sheet("With ISO 22000"),data,"iso22000",True,"Suppliers WITH ISO 22000","1A6B3A",category)
        _cert_sheet(wb.create_sheet("No ISO 22000"),data,"iso22000",False,"Suppliers WITHOUT ISO 22000","8B1A1A",category)
        _cert_sheet(wb.create_sheet("With FSSC 22000"),data,"fssc22000",True,"Suppliers WITH FSSC 22000","1A6B3A",category)
        _cert_sheet(wb.create_sheet("No FSSC 22000"),data,"fssc22000",False,"Suppliers WITHOUT FSSC 22000","8B1A1A",category)
    else:
        _cert_sheet(wb.create_sheet("With CEP"),data,"cep",True,"Suppliers WITH CEP (EDQM)","1A6B3A",category)
        _cert_sheet(wb.create_sheet("No CEP"),data,"cep",False,"Suppliers WITHOUT CEP","8B1A1A",category)
        _cert_sheet(wb.create_sheet("With GMP"),data,"gmp",True,"Suppliers WITH GMP","1A6B3A",category)
        _cert_sheet(wb.create_sheet("No GMP"),data,"gmp",False,"Suppliers WITHOUT GMP","8B1A1A",category)

    wb.save(output_path)


def _cert_sheet(ws,data,cert_key,cert_val,title,color,category):
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    def fl(c): return PatternFill("solid",fgColor=c)
    def brd():
        s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)
    label=cert_key.upper().replace("ISO22000","ISO 22000").replace("FSSC22000","FSSC 22000")
    cols=["Ingredient" if category=="bad" else "API","CAS","Supplier","Country","Contact",label]
    widths=[22,14,28,14,30,18]
    ws.merge_cells("A1:F1")
    c=ws["A1"]; c.value=title; c.font=Font(bold=True,size=12,color="FFFFFF")
    c.fill=fl(color); c.alignment=Alignment(horizontal="center",vertical="center")
    for i,(h,w) in enumerate(zip(cols,widths),1):
        c=ws.cell(row=2,column=i,value=h); c.font=Font(bold=True,color="FFFFFF",size=10)
        c.fill=fl("2E86AB"); c.border=brd(); c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(i)].width=w
    row=3
    for entry in data:
        for sup in entry.get("suppliers",[]):
            v=sup.get(cert_key)
            if cert_val is True and v is not True: continue
            if cert_val is False and v is not False: continue
            bg="F5F8FC" if row%2==0 else "FFFFFF"
            ok=v is True
            vals=[entry["name"].title(),entry.get("cas","-"),sup.get("name","-"),
                  sup.get("country","-"),sup.get("contact","-"),"YES" if ok else "NO"]
            for col,val in enumerate(vals,1):
                c=ws.cell(row=row,column=col,value=val); c.border=brd()
                c.alignment=Alignment(horizontal="center" if col in[1,2,4,6] else "left",vertical="center")
                if col==6:
                    c.fill=fl("D4EDDA" if ok else "FADADD")
                    c.font=Font(bold=True,color="1A6B3A" if ok else "8B1A1A")
                elif col==5: c.fill=fl(bg); c.font=Font(color="1155CC",size=9)
                else: c.fill=fl(bg)
            row+=1
    ws.freeze_panes="A3"


def main():
    print()
    print("="*60)
    print("  💊 API FINDER — Пошук постачальників АФІ / БАД")
    print("="*60)
    print()

    while True:
        print("  Оберіть категорію пошуку:")
        print()
        print("    1 - Фарма (АФІ)  → перевірка CEP + GMP")
        print("    2 - БАД (сировина) → перевірка ISO 22000 + FSSC 22000")
        print()

        while True:
            choice=input("  Категорія (1 або 2): ").strip()
            if choice in ("1","2"): break
            if choice.lower() in ("exit","quit"): print("👋 До побачення!"); return
            print("  ⚠️  Введіть 1 або 2")

        category="pharma" if choice=="1" else "bad"
        cat_label="Фарма (АФІ)" if category=="pharma" else "БАД (сировина)"

        print(f"\n  ✅ Режим: {cat_label}")
        if category=="bad":
            print("  📋 Сертифікати: ISO 22000, FSSC 22000, GMP (опціонально)")
        else:
            print("  📋 Сертифікати: CEP (EDQM), GMP")

        print()
        print("  Введіть субстанції (кожна на новому рядку):")
        print("    Caffeine (CAS 58-08-2)")
        print("  Порожній рядок = почати пошук")
        print()

        lines=[]
        while True:
            try: line=input("   ")
            except (EOFError,KeyboardInterrupt): print(); return
            if line.strip().lower() in ("exit","quit"): print("👋 До побачення!"); return
            if not line.strip():
                if lines: break
                continue
            lines.append(line)

        substances=parse_input("\n".join(lines))
        if not substances: print("⚠️ Не вдалося розпізнати введення.\n"); continue

        print(f"\n📋 Розпізнано {len(substances)} субстанцій ({cat_label}):\n")
        for s in substances:
            print(f"   * {s['name']}  ->  CAS: {s.get('cas','-')}")

        print(f"\n⏳ Починаю пошук...\n")

        results=[]
        for i,sub in enumerate(substances,1):
            name=sub["name"]; cas=sub.get("cas")
            print(f"[{i}/{len(substances)}] {name} (CAS {cas or '?'}):", flush=True)

            print("   📘 PubChem ...", end=" ", flush=True)
            pc=pubchem_by_cas(cas) if cas else pubchem_by_name(name)
            if not cas and pc.get("cas"): cas=pc["cas"]
            formula=pc.get("formula",""); mw=pc.get("mw","")
            print(f"formula={formula or '?'}, MW={mw or '?'}")

            print("   🤖 Пошук постачальників ...", end=" ", flush=True)
            suppliers=ai_search_suppliers(name, cas or "-", formula, category)

            if category=="bad":
                iso_c=sum(1 for s in suppliers if s.get("iso22000") is True)
                fssc_c=sum(1 for s in suppliers if s.get("fssc22000") is True)
                print(f"знайдено: {len(suppliers)} (ISO22000: {iso_c}, FSSC: {fssc_c})")
            else:
                cep_c=sum(1 for s in suppliers if s.get("cep") is True)
                gmp_c=sum(1 for s in suppliers if s.get("gmp") is True)
                print(f"знайдено: {len(suppliers)} (CEP: {cep_c}, GMP: {gmp_c})")

            if i<len(substances): time.sleep(1)
            results.append({"name":name,"cas":cas or "-","formula":formula,"mw":mw,"suppliers":suppliers})

        filename=generate_filename(substances,category)
        output_path=REPORTS_DIR/filename
        generate_excel(results,str(output_path),category)

        total_s=sum(len(e.get("suppliers",[])) for e in results)
        print(f"\n{'='*50}")
        print(f"  ✅ Режим:        {cat_label}")
        print(f"  Субстанцій:      {len(results)}")
        print(f"  Постачальників:  {total_s}")
        if category=="bad":
            iso_y=sum(1 for e in results for s in e.get("suppliers",[]) if s.get("iso22000") is True)
            fssc_y=sum(1 for e in results for s in e.get("suppliers",[]) if s.get("fssc22000") is True)
            print(f"  З ISO 22000:     {iso_y}")
            print(f"  З FSSC 22000:    {fssc_y}")
        else:
            cep_y=sum(1 for e in results for s in e.get("suppliers",[]) if s.get("cep") is True)
            gmp_y=sum(1 for e in results for s in e.get("suppliers",[]) if s.get("gmp") is True)
            print(f"  З CEP:           {cep_y}")
            print(f"  З GMP:           {gmp_y}")
        print(f"  Файл:            {output_path}")
        print(f"{'='*50}\n")

        json_path=str(output_path).replace(".xlsx",".json")
        with open(json_path,"w",encoding="utf-8") as f:
            json.dump(results,f,ensure_ascii=False,indent=2)

        print("✅ Готово! Введіть новий запит або 'exit' для виходу.\n")


if __name__=="__main__":
    main()
